from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import (
    ListView, CreateView, UpdateView, DeleteView, DetailView
)
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_http_methods
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.utils import timezone
from collections import defaultdict
import csv
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

from accounts.models import User
from .models import Questionnaire, Question, QuestionOption, Response, Answer
from .forms import QuestionnaireForm, QuestionForm, ResponseForm
from patients.models import PatientVitals

# Questionnaire Views
class QuestionnaireListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Questionnaire
    template_name = 'questionnaires/questionnaire_list.html'
    context_object_name = 'questionnaires'
    paginate_by = 10
    
    def test_func(self):
        return self.request.user.is_authenticated and (
            self.request.user.role == User.Role.SUPER_ADMIN or 
            self.request.user.is_superuser
        )

    def handle_no_permission(self):
        if self.request.user.is_authenticated:
            if self.request.user.role == User.Role.HEALTH_ASSISTANT:
                messages.warning(self.request, "Access Denied: You have been redirected to your dashboard.")
                return redirect('health_assistant:home')
            elif self.request.user.role == User.Role.DOCTOR:
                messages.warning(self.request, "Access Denied: You have been redirected to your dashboard.")
                return redirect('doctor:home')
        return super().handle_no_permission()
    
    def get_queryset(self):
        return Questionnaire.objects.all().order_by('-created_at')

class QuestionnaireCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Questionnaire
    form_class = QuestionnaireForm
    template_name = 'questionnaires/questionnaire_form.html'
    
    def test_func(self):
        return self.request.user.role == User.Role.SUPER_ADMIN
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'Questionnaire created successfully.')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse_lazy('questionnaires:detail', kwargs={'pk': self.object.pk})

class QuestionnaireUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Questionnaire
    form_class = QuestionnaireForm
    template_name = 'questionnaires/questionnaire_form.html'
    
    def test_func(self):
        return self.request.user.role == User.Role.SUPER_ADMIN
    
    def form_valid(self, form):
        messages.success(self.request, 'Questionnaire updated successfully.')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse_lazy('questionnaires:detail', kwargs={'pk': self.object.pk})

class QuestionnaireDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Questionnaire
    template_name = 'questionnaires/questionnaire_detail.html'
    context_object_name = 'questionnaire'
    
    def test_func(self):
        return self.request.user.role == User.Role.SUPER_ADMIN
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['questions'] = self.object.questions.all().order_by('order')
        return context

class QuestionnaireDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Questionnaire
    template_name = 'questionnaires/questionnaire_confirm_delete.html'
    success_url = reverse_lazy('questionnaires:list')
    
    def test_func(self):
        return self.request.user.role == User.Role.SUPER_ADMIN
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Questionnaire deleted successfully.')
        return super().delete(request, *args, **kwargs)

# Question Views
class QuestionCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Question
    form_class = QuestionForm
    template_name = 'questionnaires/question_form.html'
    
    def test_func(self):
        return self.request.user.role == User.Role.SUPER_ADMIN
    
    def get_initial(self):
        initial = super().get_initial()
        questionnaire_id = self.kwargs.get('questionnaire_id')
        if questionnaire_id:
            initial['questionnaire'] = get_object_or_404(Questionnaire, id=questionnaire_id)
        return initial
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        questionnaire_id = self.kwargs.get('questionnaire_id')
        if questionnaire_id:
            context['questionnaire'] = get_object_or_404(Questionnaire, id=questionnaire_id)
        return context
    
    def form_valid(self, form):
        questionnaire_id = self.kwargs.get('questionnaire_id')
        questionnaire = get_object_or_404(Questionnaire, id=questionnaire_id)
        form.instance.questionnaire = questionnaire
        
        # Set the display order to be the next available number
        last_question = (
            Question.objects.filter(questionnaire=questionnaire)
            .order_by('-order')
            .first()
        )
        form.instance.order = (last_question.order + 1) if last_question else 1
        
        response = super().form_valid(form)
        question = self.object
        
        # Process followups
        followups_data = self.request.POST.get('followups_data')
        if followups_data and form.instance.question_type == 'yes_no':
            import json
            try:
                followups = json.loads(followups_data)
                for index, fu in enumerate(followups):
                    Question.objects.create(
                        questionnaire=question.questionnaire,
                        parent=question,
                        trigger_answer=fu['trigger'],
                        question_text=fu['text'],
                        question_type=fu['type'],
                        is_required=fu['required'],
                        order=question.order + index + 1
                    )
            except Exception as e:
                print("Error parsing followups:", e)
                
        messages.success(self.request, 'Question added successfully.')
        return response
    
    def get_success_url(self):
        return reverse_lazy('questionnaires:detail', 
                          kwargs={'pk': self.object.questionnaire.id})

class QuestionUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Question
    form_class = QuestionForm
    template_name = 'questionnaires/question_form.html'
    
    def test_func(self):
        return self.request.user.role == User.Role.SUPER_ADMIN
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['questionnaire'] = self.object.questionnaire
        return context
    
    def form_valid(self, form):
        response = super().form_valid(form)
        question = self.object
        
        followups_data = self.request.POST.get('followups_data')
        if followups_data and question.question_type == 'yes_no':
            import json
            try:
                followups = json.loads(followups_data)
                processed_ids = []
                for index, fu in enumerate(followups):
                    fu_id = fu.get('id')
                    if fu_id:
                        try:
                            child = Question.objects.get(id=fu_id, parent=question)
                            child.question_text = fu['text']
                            child.question_type = fu['type']
                            child.is_required = fu['required']
                            child.trigger_answer = fu['trigger']
                            child.order = question.order + index + 1
                            child.save()
                            processed_ids.append(child.id)
                        except Question.DoesNotExist:
                            pass
                    else:
                        child = Question.objects.create(
                            questionnaire=question.questionnaire,
                            parent=question,
                            trigger_answer=fu['trigger'],
                            question_text=fu['text'],
                            question_type=fu['type'],
                            is_required=fu['required'],
                            order=question.order + index + 1
                        )
                        processed_ids.append(child.id)
                
                question.follow_ups.exclude(id__in=processed_ids).delete()
            except Exception as e:
                print("Error parsing followups:", e)
        elif question.question_type != 'yes_no':
            question.follow_ups.all().delete()
            
        messages.success(self.request, 'Question updated successfully.')
        return response
    
    def get_success_url(self):
        return reverse_lazy('questionnaires:detail', 
                          kwargs={'pk': self.object.questionnaire.id})

class QuestionDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Question
    template_name = 'questionnaires/question_confirm_delete.html'
    
    def test_func(self):
        return self.request.user.role == User.Role.SUPER_ADMIN
    
    def get_success_url(self):
        questionnaire_id = self.object.questionnaire.id
        return reverse_lazy('questionnaires:detail', 
                          kwargs={'pk': questionnaire_id})
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Question deleted successfully.')
        return super().delete(request, *args, **kwargs)

# Response Views
class ResponseListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Response
    template_name = 'questionnaires/response_list.html'
    context_object_name = 'responses'
    paginate_by = 20
    
    def test_func(self):
        from accounts.models import User
        return self.request.user.is_staff or self.request.user.role in [User.Role.HEALTH_ASSISTANT, User.Role.DOCTOR]
    
    def get_template_names(self):
        from accounts.models import User
        if self.request.user.role in [User.Role.HEALTH_ASSISTANT, User.Role.DOCTOR]:
            return ['health_assistant/response_list.html']
        return [self.template_name]
    
    def get_queryset(self):
        queryset = Response.objects.select_related('questionnaire', 'respondent', 'patient')
        
        # Filter by questionnaire if specified
        questionnaire_id = self.request.GET.get('questionnaire')
        if questionnaire_id:
            queryset = queryset.filter(questionnaire_id=questionnaire_id)
            
        # Filter by respondent if specified
        respondent_id = self.request.GET.get('respondent')
        if respondent_id:
            queryset = queryset.filter(respondent_id=respondent_id)
            
        # Filter by patient if specified
        patient_id = self.request.GET.get('patient')
        if patient_id:
            queryset = queryset.filter(patient__patient_id__icontains=patient_id)
            
        # Filter by date range if specified
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        
        if date_from:
            try:
                from datetime import datetime
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                queryset = queryset.filter(started_at__date__gte=date_from_obj)
            except ValueError:
                pass  # Invalid date format, ignore filter
                
        if date_to:
            try:
                from datetime import datetime
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                queryset = queryset.filter(started_at__date__lte=date_to_obj)
            except ValueError:
                pass  # Invalid date format, ignore filter
            
        return queryset.order_by('-submitted_at', '-started_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['questionnaires'] = Questionnaire.objects.filter(is_active=True)
        return context

class ResponseDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Response
    template_name = 'questionnaires/response_detail.html'
    context_object_name = 'response'
    
    def test_func(self):
        from accounts.models import User
        if self.request.user.is_staff or self.request.user.role in [User.Role.HEALTH_ASSISTANT, User.Role.DOCTOR]:
            return True
        return self.request.user == self.get_object().respondent
    
    def get_template_names(self):
        from accounts.models import User
        if self.request.user.role in [User.Role.HEALTH_ASSISTANT, User.Role.DOCTOR]:
            return ['health_assistant/response_detail.html']
        return [self.template_name]
    
    def get_queryset(self):
        return Response.objects.select_related('questionnaire', 'respondent', 'patient')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['answers'] = self.object.answers.select_related('question')
        
        # Use the specifically linked vital snapshot if it exists
        if self.object.vitals:
            context['vitals'] = self.object.vitals
        # Fallback to vitals recorded around the time of the response (for older records)
        elif self.object.patient:
            from patients.models import PatientVitals
            # Prefer vitals recorded before or at the time of submission
            from django.utils import timezone
            base_time = self.object.submitted_at or self.object.started_at or timezone.now()
            
            # Find the most recent vitals recorded *before* this response was submitted
            context['vitals'] = PatientVitals.objects.filter(
                patient=self.object.patient,
                recorded_at__lte=base_time
            ).order_by('-recorded_at').first()
            
            # If no vitals were recorded before (e.g. recorded slightly after or same time),
            # just get the very latest as the best guess
            if not context['vitals']:
                context['vitals'] = PatientVitals.objects.filter(
                    patient=self.object.patient
                ).order_by('-recorded_at').first()
        else:
            context['vitals'] = None
        return context

class ResponseDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Response
    template_name = 'questionnaires/response_confirm_delete.html'
    
    def test_func(self):
        # Only staff or the respondent can delete (if respondent is not just a health assistant reviewing, but actual admin)
        # Actually requirement says "not delete" for health assistant, so keep just staff or original respondent
        return self.request.user.is_staff
    
    def get_success_url(self):
        return reverse_lazy('questionnaires:response_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Response deleted successfully.')
        return super().delete(request, *args, **kwargs)


@login_required
def get_response_edit_form(request, pk):
    """Returns a partial HTML form for editing a response with permission check."""
    try:
        from accounts.models import User
        response_obj = get_object_or_404(Response, pk=pk)

        # Permission check: Staff, Doctors, and Health Assistants can edit. Others can only edit their own.
        has_permission = (
            request.user.is_staff or 
            request.user.role in [User.Role.HEALTH_ASSISTANT, User.Role.DOCTOR] or
            request.user == response_obj.respondent
        )
        
        if not has_permission:
            return JsonResponse({
                'success': False, 
                'message': 'Permission Denied: You do not have authority to edit this record.'
            }, status=403)

        # Get all questions in the questionnaire, not just the answered ones
        questions = response_obj.questionnaire.questions.all().order_by('order', 'id')
        
        # If a specific question ID is provided, figure out exactly which questions to show 
        # (the question itself, plus any conditional descendants)
        target_question_id = request.GET.get('question_id')
        if target_question_id:
            try:
                target_question_id = int(target_question_id)
                target_question = questions.get(id=target_question_id)
                # Keep the target question and all its descendants
                descendants = target_question.get_all_descendants()
                descendant_ids = [d.id for d in descendants]
                allowed_ids = set([target_question_id] + descendant_ids)
                questions = [q for q in questions if q.id in allowed_ids]
            except (ValueError, Question.DoesNotExist):
                pass
                
        # Create a map of question_id to answer for easier lookup
        answers_map = {a.question_id: a for a in response_obj.answers.all()}
        
        # Bundle question and answer together for easy template iteration
        bundled_data = []
        allowed_ids = {q.id for q in questions}
        for q in questions:
            is_parent_present = bool(q.parent_id and q.parent_id in allowed_ids)
            bundled_data.append({
                'question': q,
                'answer': answers_map.get(q.id),
                'is_hidden': bool(q.parent and is_parent_present)
            })
        
        return render(request, 'questionnaires/partials/response_edit_form.html', {
            'response': response_obj,
            'bundled_data': bundled_data
        })
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'message': f'Server Error: {str(e)}'
        }, status=500)


@login_required
@require_POST
def api_update_response(request, pk):
    """AJAX endpoint — saves edited answers from the response detail modal."""
    import json
    response_obj = get_object_or_404(Response, pk=pk)

    if not (request.user.is_staff or request.user.role in ['HEALTH_ASSISTANT', 'DOCTOR']):
        return JsonResponse({'success': False, 'message': 'Access denied'}, status=403)

    try:
        data = json.loads(request.body)
        answers_data = data.get('answers', {})  # {question_id: value}

        for question_id_str, value in answers_data.items():
            try:
                from questionnaires.models import Answer, Question, QuestionOption
                question = Question.objects.get(pk=int(question_id_str))
                answer, _ = Answer.objects.get_or_create(response=response_obj, question=question)

                q_type = question.question_type
                if q_type in ('short_answer', 'long_answer', 'yes_no', 'true_false', 'number'):
                    answer.text_answer = str(value) if value is not None else ''
                    answer.option_answer.clear()
                elif q_type in ('multiple_choice', 'image_choice'):
                    answer.text_answer = ''
                    answer.option_answer.clear()
                    if value is not None:
                        # Convert all values to integers safely
                        ids = []
                        if isinstance(value, list):
                            for v in value:
                                try:
                                    if v: ids.append(int(v))
                                except (ValueError, TypeError):
                                    continue
                        else:
                            try:
                                if value: ids.append(int(value))
                            except (ValueError, TypeError):
                                pass
                        
                        if ids:
                            options = QuestionOption.objects.filter(pk__in=ids)
                            answer.option_answer.set(options)
                elif q_type == 'date':
                    answer.text_answer = str(value) if value else ''
                answer.save()
            except (Question.DoesNotExist, ValueError):
                continue

        return JsonResponse({'success': True, 'message': 'Response updated successfully!'})

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)



@login_required
@require_http_methods(['POST'])
def update_question_order(request):
    try:
        question_order = request.POST.getlist('order[]')
        with transaction.atomic():
            for index, question_id in enumerate(question_order, start=1):
                Question.objects.filter(id=question_id).update(order=index)
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@login_required
def api_list_questionnaires(request):
    """API endpoint to list available questionnaires"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    questionnaires = Questionnaire.objects.filter(is_active=True).order_by('title')
    
    questionnaire_data = []
    for q in questionnaires:
        questionnaire_data.append({
            'id': q.id,
            'title': q.title,
            'description': q.description,
            'question_count': q.questions.count()
        })
    
    return JsonResponse({'questionnaires': questionnaire_data})

# Public Views
@login_required
def questionnaire_start(request, pk):
    questionnaire = get_object_or_404(Questionnaire, pk=pk, is_active=True)
    
    if request.method == 'POST':
        print(f"DEBUG: POST data: {request.POST}")
        print(f"DEBUG: FILES data: {request.FILES}")
        form = ResponseForm(questionnaire, request.POST, request.FILES)
        if form.is_valid():
            response = form.save(commit=False)
            response.questionnaire = questionnaire
            response.is_complete = True
            response.submitted_at = timezone.now()
            if request.user.is_authenticated:
                response.respondent = request.user
            
            # Handle patient association
            patient_id = request.POST.get('patient_id')
            if patient_id:
                from patients.models import Patient
                try:
                    patient = Patient.objects.get(id=patient_id)
                    response.patient = patient
                except Patient.DoesNotExist:
                    pass
            
            response.save()
            form.save_answers(response)
            
            # Return JSON response for AJAX requests
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Questionnaire submitted successfully!',
                    'response_id': response.pk
                })
            else:
                return redirect('questionnaires:questionnaire_thank_you', pk=response.pk)
        else:
            # Return JSON response for AJAX requests with errors
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': 'Please correct the errors below.',
                    'errors': form.errors
                })
    else:
        form = ResponseForm(questionnaire)
    
    # Use appropriate template based on questionnaire type
    if questionnaire.title.lower() == 'patient registration' or 'patient' in questionnaire.title.lower():
        template = 'questionnaires/patient_profile_form.html'
    elif questionnaire.title.lower() == 'medical screening questionnaire' or 'medical screening' in questionnaire.title.lower():
        template = 'questionnaires/simple_questionnaire_display.html'
    else:
        template = 'questionnaires/simple_questionnaire_display.html'
    
    return render(request, template, {
        'questionnaire': questionnaire,
        'questions': questionnaire.questions.all().order_by('order'),
        'form': form,
    })

def questionnaire_thank_you(request, pk):
    response = get_object_or_404(Response, pk=pk)
    return render(request, 'questionnaires/thank_you.html', {
        'response': response,
    })


@login_required
@require_POST
@csrf_exempt
def update_question_order(request):
    """API endpoint to update question order."""
    try:
        import json
        data = json.loads(request.body)
        question_ids = data.get('question_ids', [])
        
        if not question_ids:
            return JsonResponse({
                'success': False,
                'error': 'No question IDs provided'
            }, status=400)
        
        with transaction.atomic():
            for index, question_id in enumerate(question_ids, start=1):
                Question.objects.filter(id=question_id).update(order=index)
        
        return JsonResponse({
            'success': True,
            'message': 'Question order updated successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def simple_questionnaire_builder(request):
    """Simple questionnaire builder for creating medical screening questionnaires (Admin only)."""
    if request.user.role != User.Role.SUPER_ADMIN:
        messages.error(request, "Permission Denied: Only Super Admins can access the builder.")
        return redirect('dashboard:dashboard')  # Consistent with dashboard/urls.py
    return render(request, 'questionnaires/simple_questionnaire_builder.html')


@require_POST
@login_required
def upload_attachment(request):
    """Handle file upload for attachment questions."""
    try:
        if 'fileToUpload' not in request.FILES:
            return JsonResponse({
                'success': False,
                'error': 'No file uploaded'
            }, status=400)
        
        file = request.FILES['fileToUpload']
        question_id = request.POST.get('question_id')
        
        # Validate file size (10MB limit)
        max_size = 10 * 1024 * 1024  # 10MB in bytes
        if file.size > max_size:
            return JsonResponse({
                'success': False,
                'error': f'File "{file.name}" is too large ({file.size/1024/1024:.1f}MB). Maximum size is 10MB.'
            }, status=400)
        
        # Validate file type
        allowed_types = [
            'application/pdf', 'application/vnd.ms-excel', 
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'text/csv', 'text/plain', 'application/msword',
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            'image/jpeg', 'image/png', 'image/gif', 'image/bmp'
        ]
        
        if file.content_type not in allowed_types:
            return JsonResponse({
                'success': False,
                'error': f'File "{file.name}" has invalid format. Allowed: PDF, XLS, XLSX, CSV, TXT, DOC, DOCX, JPG, PNG, GIF, BMP'
            }, status=400)
        
        # Store file temporarily (will be saved when form is submitted)
        # For now, just return success with file info
        file_size_mb = file.size / 1024 / 1024
        
        return JsonResponse({
            'success': True,
            'message': f'File "{file.name}" ({file_size_mb:.2f}MB) uploaded successfully',
            'filename': file.name,
            'size': file_size_mb
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Upload failed: {str(e)}'
        }, status=500)


@login_required
def download_responses(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    from patients.models import PatientVitals
    from collections import defaultdict
    import io

    # Get filter parameters
    questionnaire_id = request.GET.get('questionnaire')
    patient_id = request.GET.get('patient')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Filter responses
    responses = Response.objects.all().select_related('questionnaire', 'patient', 'respondent').prefetch_related('answers__question', 'answers__option_answer')
    
    if questionnaire_id:
        responses = responses.filter(questionnaire_id=questionnaire_id)
    if patient_id:
        responses = responses.filter(patient__patient_id__icontains=patient_id)
    
    # Filter by date range if specified
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            responses = responses.filter(started_at__date__gte=date_from_obj)
        except ValueError:
            pass
            
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            responses = responses.filter(started_at__date__lte=date_to_obj)
        except ValueError:
            pass
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Group responses by questionnaire
    questionnaire_responses = defaultdict(list)
    for response in responses:
        questionnaire_responses[response.questionnaire].append(response)
        
    if not questionnaire_responses:
        # Empty state
        ws = wb.create_sheet(title="No Responses")
        ws.append(["No data available for the given filters."])
    
    # Define headers for vitals
    vitals_headers = [
        'BP (Systolic/Diastolic)', 'Heart Rate (bpm)', 'Respiratory Rate/min', 
        'Temperature (°C)', 'SpO2 (%)', 'Weight (kg)', 'Height (cm)', 'BMI'
    ]
    
    for questionnaire, q_responses in questionnaire_responses.items():
        # Excel sheet names have a 31-char limit and don't allow some special chars
        safe_title = "".join([c if c.isalnum() else " " for c in questionnaire.title])[:31].strip()
        if not safe_title:
            safe_title = f"Questionnaire_{questionnaire.id}"
            
        sheet_name = safe_title
        counter = 1
        while sheet_name in wb.sheetnames:
            suffix = f"_{counter}"
            avail_len = 31 - len(suffix)
            sheet_name = safe_title[:avail_len] + suffix
            counter += 1
            
        ws = wb.create_sheet(title=sheet_name)
        
        # Write header
        header = [
            'Response ID', 'Patient ID', 'Patient Name', 'Questionnaire', 
            'Respondent', 'Status', 'Started At'
        ]
        
        header.extend(vitals_headers)
        
        questions = questionnaire.questions.all().order_by('order')
        for question in questions:
            header.append(f'Q{question.get_display_number()}: {question.question_text[:50]}...')
            
        ws.append(header)
        
        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data rows
        for response in q_responses:
            # Use specifically linked vitals snapshot if available
            vitals = response.vitals
            if not vitals and response.patient:
                # Fallback: Find vitals recorded at or before the response time
                # Uses submitted_at (if complete) or started_at
                ref_time = response.submitted_at or response.started_at or timezone.now()
                vitals = PatientVitals.objects.filter(
                    patient=response.patient,
                    recorded_at__lte=ref_time
                ).order_by('-recorded_at').first()
                
                # If still none recorded before (backwards compatibility), get very latest
                if not vitals:
                    vitals = PatientVitals.objects.filter(
                        patient=response.patient
                    ).order_by('-recorded_at').first()
            
            row = [
                response.id,
                response.patient.patient_id if response.patient else '',
                f"{response.patient.first_name} {response.patient.last_name}" if response.patient else '',
                response.questionnaire.title,
                response.respondent.get_full_name() if response.respondent else '',
                'Complete' if response.is_complete else 'In Progress',
                response.started_at.strftime('%Y-%m-%d %H:%M:%S') if response.started_at else ''
            ]
            
            if vitals:
                bp = f"{vitals.blood_pressure_systolic or '-'}/{vitals.blood_pressure_diastolic or '-'}" if (vitals.blood_pressure_systolic or vitals.blood_pressure_diastolic) else "-"
                row.extend([
                    bp,
                    vitals.heart_rate or '-',
                    vitals.respiratory_rate or '-',
                    vitals.temperature or '-',
                    vitals.spo2 or '-',
                    vitals.weight or '-',
                    vitals.height or '-',
                    vitals.bmi or '-'
                ])
            else:
                row.extend(['-'] * len(vitals_headers))
            
            answers_dict = {answer.question_id: answer for answer in response.answers.all()}
            for question in questions:
                answer = answers_dict.get(question.id)
                if answer:
                    if answer.file_answer:
                        file_url = request.build_absolute_uri(answer.file_answer.url)
                        row.append(f'=HYPERLINK("{file_url}", "{answer.file_answer.name}")')
                    elif answer.option_answer.exists():
                        options_text = []
                        for opt in answer.option_answer.all():
                            opt_text = opt.text or ''
                            if getattr(opt, 'option_image', None) and not opt_text:
                                opt_text = '[Image option]'
                            if getattr(opt, 'option_image', None) and opt.option_image:
                                file_url = request.build_absolute_uri(opt.option_image.url) if hasattr(request, 'build_absolute_uri') else opt.option_image.url
                                opt_text += f" (Image URL: {file_url})"
                            options_text.append(opt_text.strip())
                        
                        options = ', '.join(options_text)
                        row.append(options)
                    else:
                        row.append(answer.text_answer or '')
                else:
                    row.append('')
                    
            ws.append(row)
            
        # Adjust column widths
        for col_idx, column_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), 1):
            if col_idx > len(header):
                break
            length = max((len(str(cell.value) if cell.value is not None else "")) for cell in column_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(length + 2, 50)
            
    # Output file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return Excel response
    response = HttpResponse(
        output.read(), 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"questionnaire_responses_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response
